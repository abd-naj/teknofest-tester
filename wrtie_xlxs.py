class wirteToXlxs:
    import pandas as pd
    import openpyxl
    import time
    filName = ''
    # df = ''
    path = ''
    dataToWrite = [[0, 0, 0, 0]]
    def __init__(self, filName):
        self.filName = filName
        print('dddd')
        df = self.pd.DataFrame(self.dataToWrite, columns=['LD1', 'LD2', 'speed', 'time'])
        # print(df)
        df.to_excel(self.filName, sheet_name = 'sheet1')
        self.path = self.filName

    def write(self, dataLd1, dataLd2, speed, time):
        print('fffffff')
        # print(dataLd1)
        # print(dataLd2)
        # print(speed)
        # data = list(zip(self.dataToWrite, [dataLd1, dataLd2, speed, 0]))
        self.dataToWrite.append([dataLd1, dataLd2, speed, time])
        print(self.dataToWrite)
        # with self.pd.ExcelWriter(self.filName) as writer:
        #     self.df.to_excel(writer, sheet_name='sheet1')

        df = self.pd.DataFrame(self.dataToWrite, columns=['LD1', 'LD2', 'speed', 'time'])
        print('df', df)

        with self.pd.ExcelWriter(self.path, engine="openpyxl", mode='a') as writer:
            # writer.book = self.openpyxl.load_workbook(self.path)
            df.to_excel(writer)

         ###################################
        # book = self.openpyxl.load_workbook(self.path)
        # writer = self.pd.ExcelWriter(self.path, engine='openpyxl')
        # writer.book = book
        # writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        #
        # data_filtered.to_excel(writer, "Main", cols=['Diff1', 'Diff2'])

         #########################


        # self.filName = self.filName
        # now = self.time.time()
        # # create excel file
        # # Create a workbook and add a worksheet.
        # fileName = self.filName
        # workbook = self.xlsxwriter.Workbook(fileName)
        # worksheet = workbook.add_worksheet()
        # # Some data we want to write to the worksheet.
        # expenses = (
        #     [1000, 111, 1222, now],
        #     [2000, 222, 2222, now],
        #     [3000, 333, 3222, now],
        #     [4000, 444, 4222, now],
        #     [5000, 444, 4222, now],
        # )







        # Iterate over the data and write it out row by row.
        # for LD1, LD2, speed, time in (expenses):
        #     worksheet.write(row, col,     LD1)
        #     worksheet.write(row, col + 1, LD2)
        #     worksheet.write(row, col + 2, speed)
        #     worksheet.write(row, col + 3, time)
        #     row += 1

        # Write a total using a formula.
        # worksheet.write(row, 0, 'Total')
        # worksheet.write(row, 1, '=SUM(B1:B4)')

        # workbook.close()